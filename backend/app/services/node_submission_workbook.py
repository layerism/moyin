import json
import re
from datetime import UTC, datetime
from io import BytesIO
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.domain.form_fields import OTHER_OPTION_ID, normalize_form_fields
from app.domain.workflow import confirmation_requires_scans
from app.repositories.teacher_node_exports import (
    TeacherNodeExportSelection,
    TeacherNodeExportStudent,
)

_UNSAFE_FILENAME_CHARACTERS = re.compile(r"[\\/\x00-\x1f\x7f]")
_SHANGHAI = ZoneInfo("Asia/Shanghai")
_SUBMISSION_STATUS_LABELS = {
    "reviewing": "审核中",
    "approved": "已通过",
    "rejected": "未通过",
    "audit_error": "审核异常",
}
_AUDIT_STATUS_LABELS = {
    "pending": "等待审核",
    "running": "审核中",
    "succeeded": "审核完成",
    "failed": "审核异常",
    "cancelled": "已取消",
}


def _safe_filename_component(value: str) -> str:
    component = _UNSAFE_FILENAME_CHARACTERS.sub("_", value).replace("..", "_").strip(" .")
    return component or "未命名"


def _excel_datetime(value: str | None) -> datetime | None:
    if not value:
        return None
    parsed = datetime.fromisoformat(value)
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=UTC)
    return parsed.astimezone(_SHANGHAI).replace(tzinfo=None)


def _excel_cell(value: object) -> object:
    if isinstance(value, str) and value[:1] in {"=", "+", "-", "@", "\t", "\r"}:
        return f"'{value}"
    return value


def _format_form_answer(field: dict[str, object], raw_value: object) -> str:
    if field.get("type") in {"text", "textarea"}:
        return raw_value if isinstance(raw_value, str) else ""
    if not isinstance(raw_value, dict):
        return ""

    option_labels = {
        str(option["id"]): str(option["label"])
        for option in field.get("options", [])
        if isinstance(option, dict) and "id" in option and "label" in option
    }
    other_text = raw_value.get("otherText")

    def label(option_id: str) -> str:
        if option_id == OTHER_OPTION_ID:
            return f"其他：{other_text}" if isinstance(other_text, str) and other_text else "其他"
        return option_labels.get(option_id, option_id)

    if field.get("type") == "radio":
        selected = raw_value.get("selectedOptionId")
        return label(selected) if isinstance(selected, str) else ""

    selected = raw_value.get("selectedOptionIds")
    if not isinstance(selected, list):
        return ""
    return "；".join(label(value) for value in selected if isinstance(value, str))


def _audit_values(student: TeacherNodeExportStudent, node: dict[str, object]) -> list[object]:
    if student.submission_status is None:
        return [None, None, None, None, None]
    result = student.audit_result
    details = result.get("details") if isinstance(result.get("details"), dict) else {}
    mode = student.audit_params.get("scanAuditMode") or node.get("scanAuditMode")
    review_source = details.get("reviewSource")
    audit_method = (
        "人工审核"
        if review_source == "manual"
        else "AI 评分"
        if mode == "score"
        else "AI 通过 / 不通过"
    )
    score = details.get("score")
    passed = result.get("passed")
    conclusion: object = None
    audit_score: object = None
    if review_source == "manual" and isinstance(passed, bool):
        conclusion = "通过" if passed else "不通过"
    elif mode == "score":
        if not isinstance(score, bool) and isinstance(score, (int, float)):
            audit_score = score
        if isinstance(passed, bool):
            conclusion = "通过" if passed else "不通过"
    elif isinstance(passed, bool):
        conclusion = "通过" if passed else "不通过"
    reason = result.get("reason")
    return [
        _AUDIT_STATUS_LABELS.get(student.audit_job_status or "", student.audit_job_status),
        audit_method,
        conclusion,
        audit_score,
        reason if isinstance(reason, str) else None,
    ]


def _headers(selection: TeacherNodeExportSelection) -> list[str]:
    node = selection.node
    headers = ["序号", "学号", "姓名", "提交状态", "提交时间"]
    kind = node.get("kind")
    if kind == "form":
        headers.extend(
            str(field["label"])
            for field in normalize_form_fields(node.get("infoFields", []))
        )
    elif kind == "file":
        headers.extend(["文件名", "文件大小（MB）", "文件类型"])
    elif kind in {"announcement", "confirmation"}:
        headers.append("是否确认")
        if confirmation_requires_scans(node):
            if node.get("templateAsset") is not None:
                headers.append("模板下载时间")
            headers.extend(["扫描件数量", "扫描件名称", "扫描件页数"])
    elif kind == "answer_sheet":
        for index, _question in enumerate(node.get("answerSheet", {}).get("questions", []), start=1):
            headers.extend([f"第{index}题作答", f"第{index}题得分"])
        headers.extend(["总分", "满分", "是否及格"])
    if isinstance(node.get("auditScriptId"), str) and node.get("auditScriptId"):
        headers.extend(["审核状态", "审核方式", "审核结论", "审核分数", "审核说明"])
    return headers


def _row_values(
    selection: TeacherNodeExportSelection,
    student: TeacherNodeExportStudent,
    index: int,
) -> list[object]:
    node = selection.node
    submitted = student.submission_status is not None
    values: list[object] = [
        index,
        student.student_no,
        student.name,
        _SUBMISSION_STATUS_LABELS.get(student.submission_status or "", student.submission_status),
        _excel_datetime(student.submitted_at),
    ]
    kind = node.get("kind")
    if kind == "form":
        values.extend(
            (
                _format_form_answer(field, student.payload.get(str(field["answerKey"])))
                if submitted
                else None
            )
            for field in normalize_form_fields(node.get("infoFields", []))
        )
    elif kind == "file":
        file = student.files[0] if submitted and student.files else None
        values.extend(
            [
                file.original_name if file else None,
                round(file.size_bytes / 1024 / 1024, 3) if file else None,
                file.content_type if file else None,
            ]
        )
    elif kind in {"announcement", "confirmation"}:
        values.append("是" if submitted and student.payload.get("confirmed") is True else None)
        if confirmation_requires_scans(node):
            if node.get("templateAsset") is not None:
                values.append(_excel_datetime(student.template_downloaded_at))
            values.extend(
                [
                    len(student.files) if submitted else None,
                    "\n".join(file.original_name for file in student.files) if submitted else None,
                    (
                        "\n".join(str(file.page_count) for file in student.files)
                        if submitted
                        else None
                    ),
                ]
            )
    elif kind == "answer_sheet":
        answers = student.payload.get("answers") if submitted else None
        answers = answers if isinstance(answers, dict) else {}
        results = student.grade.get("questionResults")
        results_by_id = {
            str(result.get("questionId")): result
            for result in results
            if isinstance(result, dict) and result.get("questionId")
        } if isinstance(results, list) else {}
        for question in node.get("answerSheet", {}).get("questions", []):
            if not isinstance(question, dict):
                continue
            question_id = str(question.get("id") or "")
            answer = answers.get(question_id)
            values.extend([
                _format_answer_sheet_answer(question, answer),
                results_by_id.get(question_id, {}).get("awardedPoints") if submitted else None,
            ])
        values.extend([
            student.grade.get("score") if submitted else None,
            student.grade.get("maxScore") if submitted else None,
            (
                "是" if student.grade.get("passed") is True else "否"
                if student.grade.get("passed") is False else None
            ) if submitted else None,
        ])
    if isinstance(node.get("auditScriptId"), str) and node.get("auditScriptId"):
        values.extend(_audit_values(student, node))
    return values


def _format_answer_sheet_answer(question: dict[str, object], answer: object) -> str:
    if not isinstance(answer, dict):
        return ""
    option_labels = {
        str(option.get("id")): str(option.get("content") or "")
        for option in question.get("options", [])
        if isinstance(option, dict) and option.get("id")
    }
    if question.get("type") == "single_choice":
        selected = answer.get("selectedOptionId")
        return option_labels.get(str(selected), "") if selected else ""
    if question.get("type") == "multiple_choice":
        selected = answer.get("selectedOptionIds")
        return "；".join(
            option_labels.get(str(value), str(value))
            for value in selected
        ) if isinstance(selected, list) else ""
    if question.get("format") == "single_markdown_exact":
        answer_markdown = answer.get("answerMarkdown")
        return answer_markdown if isinstance(answer_markdown, str) else ""
    values = answer.get("blankValues")
    return "；".join(
        f"{blank.get('id')}：{values.get(str(blank.get('id')), '')}"
        for blank in question.get("blanks", [])
        if isinstance(blank, dict)
    ) if isinstance(values, dict) else ""


def build_node_submission_workbook(
    selection: TeacherNodeExportSelection,
) -> tuple[bytes, str]:
    workbook = Workbook()
    if selection.node.get("kind") == "answer_sheet":
        _build_answer_sheet_sheets(workbook, selection)
    else:
        sheet = workbook.active
        sheet.title = "节点填写数据"
        headers = _headers(selection)
        sheet.append([_excel_cell(header) for header in headers])
        for index, student in enumerate(selection.students, start=1):
            sheet.append(
                [_excel_cell(value) for value in _row_values(selection, student, index)]
            )
        _format_sheet(sheet, headers)

    output = BytesIO()
    workbook.save(output)
    node_title = str(selection.node.get("title") or selection.node.get("id") or "节点")
    filename = (
        f"{_safe_filename_component(selection.flow_name)}-"
        f"{_safe_filename_component(node_title)}-填写数据.xlsx"
    )
    return output.getvalue(), filename


def _build_answer_sheet_sheets(
    workbook: Workbook, selection: TeacherNodeExportSelection
) -> None:
    questions = [
        question for question in selection.node.get("answerSheet", {}).get("questions", [])
        if isinstance(question, dict)
    ]
    sheet = workbook.active
    sheet.title = "学生成绩"
    grade_headers = ["序号", "学号", "姓名", "提交状态", "提交时间", "提交次数"]
    grade_headers.extend(f"第{index}题得分" for index in range(1, len(questions) + 1))
    grade_headers.extend(["总分", "满分", "是否及格"])
    sheet.append(grade_headers)
    for index, student in enumerate(selection.students, start=1):
        submitted = student.submission_status is not None
        results = student.grade.get("questionResults")
        result_by_id = {
            str(result.get("questionId")): result
            for result in results
            if isinstance(result, dict) and result.get("questionId")
        } if isinstance(results, list) else {}
        values: list[object] = [
            index,
            student.student_no,
            student.name,
            _SUBMISSION_STATUS_LABELS.get(student.submission_status or "", student.submission_status),
            _excel_datetime(student.submitted_at),
            student.attempt_no if submitted else None,
        ]
        values.extend(
            result_by_id.get(str(question.get("id")), {}).get("awardedPoints")
            if submitted else None
            for question in questions
        )
        values.extend([
            student.grade.get("score") if submitted else None,
            student.grade.get("maxScore") if submitted else None,
            (
                "是" if student.grade.get("passed") is True else "否"
                if student.grade.get("passed") is False else None
            ) if submitted else None,
        ])
        sheet.append([_excel_cell(value) for value in values])
    _format_sheet(sheet, grade_headers)

    answer_sheet = workbook.create_sheet("学生答案")
    answer_headers = ["序号", "学号", "姓名", "提交状态", "提交时间"]
    answer_headers.extend(f"第{index}题作答" for index in range(1, len(questions) + 1))
    answer_sheet.append(answer_headers)
    for index, student in enumerate(selection.students, start=1):
        submitted = student.submission_status is not None
        answers = student.payload.get("answers") if submitted else None
        answers = answers if isinstance(answers, dict) else {}
        values = [
            index,
            student.student_no,
            student.name,
            _SUBMISSION_STATUS_LABELS.get(student.submission_status or "", student.submission_status),
            _excel_datetime(student.submitted_at),
            *(
                _format_answer_sheet_answer(question, answers.get(str(question.get("id"))))
                if submitted else None
                for question in questions
            ),
        ]
        answer_sheet.append([_excel_cell(value) for value in values])
    _format_sheet(answer_sheet, answer_headers)

    question_sheet = workbook.create_sheet("题目说明")
    question_headers = ["题号", "题型", "题干 Markdown", "分值", "选项", "标准答案"]
    question_sheet.append(question_headers)
    private_answers = (selection.answer_key or {}).get("answers", {})
    for index, question in enumerate(questions, start=1):
        options = question.get("options")
        option_text = "\n".join(
            f"{option.get('id')}：{option.get('content')}"
            for option in options
            if isinstance(option, dict)
        ) if isinstance(options, list) else None
        question_sheet.append([
            index,
            {"single_choice": "单选题", "multiple_choice": "多选题", "fill_blank": "填空题"}.get(str(question.get("type")), ""),
            _excel_cell(question.get("content")),
            _question_points(question),
            _excel_cell(option_text),
            _excel_cell(_format_answer_sheet_standard(question, private_answers.get(str(question.get("id"))))),
        ])
    _format_sheet(question_sheet, question_headers)


def _format_sheet(sheet, headers: list[str]) -> None:

    header_fill = PatternFill("solid", fgColor="DCE8FA")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    date_columns = {
        index
        for index, header in enumerate(headers, start=1)
        if header in {"提交时间", "模板下载时间"}
    }
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.column in date_columns and cell.value is not None:
                cell.number_format = "yyyy-mm-dd hh:mm:ss"
        row[1].number_format = "@"

    for column_index, header in enumerate(headers, start=1):
        values = [
            header,
            *(
                sheet.cell(row=row, column=column_index).value
                for row in range(2, sheet.max_row + 1)
            ),
        ]
        width = max(len(str(value)) for value in values if value is not None)
        sheet.column_dimensions[get_column_letter(column_index)].width = min(max(width + 2, 10), 40)

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.row_dimensions[1].height = 24



def _question_points(question: dict[str, object]) -> int:
    if question.get("type") == "fill_blank":
        if question.get("format") == "single_markdown_exact":
            return int(question.get("points") or 0)
        return sum(
            int(blank.get("points") or 0)
            for blank in question.get("blanks", [])
            if isinstance(blank, dict)
        )
    return int(question.get("points") or 0)


def _format_answer_sheet_standard(question: dict[str, object], answer: object) -> str:
    if not isinstance(answer, dict):
        return ""
    option_labels = {
        str(option.get("id")): str(option.get("content") or "")
        for option in question.get("options", [])
        if isinstance(option, dict) and option.get("id")
    }
    if question.get("type") == "single_choice":
        option_id = answer.get("correctOptionId")
        return option_labels.get(str(option_id), str(option_id or ""))
    if question.get("type") == "multiple_choice":
        option_ids = answer.get("correctOptionIds")
        return "；".join(
            option_labels.get(str(value), str(value))
            for value in option_ids
        ) if isinstance(option_ids, list) else ""
    if question.get("format") == "single_markdown_exact":
        accepted = answer.get("acceptedAnswerMarkdowns")
        if isinstance(accepted, list):
            return json.dumps(accepted, ensure_ascii=False, separators=(",", ":"))
        answer_markdown = answer.get("answerMarkdown")
        return answer_markdown if isinstance(answer_markdown, str) else ""
    blanks = answer.get("blanks")
    return "；".join(
        f"{blank_id}：{' / '.join(str(value) for value in details.get('acceptedAnswers', []))}"
        for blank_id, details in blanks.items()
        if isinstance(details, dict)
    ) if isinstance(blanks, dict) else ""
